import os
import glob
import openpyxl
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter

# Tenta importar o win32com para conversão em PDF (funciona somente no Windows)
try:
    import win32com.client as win32
except ImportError:
    win32 = None

# Define as pastas de entrada e saída
pasta_input = "input"
pasta_output = "output"
os.makedirs(pasta_output, exist_ok=True)

padrao_arquivos = os.path.join(pasta_input, "*.xlsx")
arquivos = glob.glob(padrao_arquivos)

if not arquivos:
    print("Nenhum arquivo .xlsx encontrado na pasta 'input'.")
else:
    for arquivo in arquivos:
        print(f"Processando arquivo: {arquivo}")
        
        # Carrega o arquivo Excel
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        ultima_linha = ws.max_row
        
        # Limpa o conteúdo das colunas C até Z (colunas 3 a 26)
        for row in ws.iter_rows(min_row=1, max_row=ultima_linha, min_col=3, max_col=26):
            for cell in row:
                cell.value = None
        
        # Aplica borda média no intervalo de C até Z
        borda_media = Side(style='medium')
        borda = Border(left=borda_media, right=borda_media, top=borda_media, bottom=borda_media)
        for row in ws.iter_rows(min_row=1, max_row=ultima_linha, min_col=3, max_col=26):
            for cell in row:
                cell.border = borda
        
        # Configura a orientação da página para paisagem
        ws.page_setup.orientation = 'landscape'
        
        # Altera a fonte de todas as células para tamanho 12
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(size=12)
        
        # Define a largura das colunas de C até Z para 4
        for col in range(3, 27):  # colunas 3 (C) até 26 (Z)
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 4
        
        # Exclui (deleta) todas as colunas após a coluna Z
        if ws.max_column > 26:
            ws.delete_cols(27, ws.max_column - 26)
        
        # Autoajusta a coluna B (segunda coluna) de acordo com o maior texto encontrado nela
        max_length = 0
        for cell in ws['B']:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions['B'].width = max_length + 2  # adiciona margem
        
        # Determina a última linha que contém dados nas colunas A ou B para definir a área de impressão
        last_print_row = 0
        for i in range(1, ws.max_row + 1):
            if ws.cell(row=i, column=1).value or ws.cell(row=i, column=2).value:
                last_print_row = i
        if last_print_row == 0:
            last_print_row = 1
        ws.print_area = f"A1:Z{last_print_row}"
        
        # Gera um novo nome para o arquivo e salva na pasta "output"
        diretorio, nome_arquivo = os.path.split(arquivo)
        nome_base, extensao = os.path.splitext(nome_arquivo)
        novo_nome = os.path.join(pasta_output, f"{nome_base}_ajustado{extensao}")
        wb.save(novo_nome)
        print(f"Arquivo salvo: {novo_nome}")
        
        # Conversão para PDF via win32com (apenas se disponível e em ambiente Windows)
        if win32 is not None:
            try:
                excel = win32.Dispatch("Excel.Application")
                excel.Visible = False
                wb_pdf = excel.Workbooks.Open(os.path.abspath(novo_nome))
                ws_pdf = wb_pdf.Worksheets[0]
                
                # Define a área de impressão para evitar espaços em branco
                ws_pdf.PageSetup.PrintArea = f"A1:Z{last_print_row}"
                # Configura para redimensionar para caber em uma única página
                ws_pdf.PageSetup.Zoom = False
                ws_pdf.PageSetup.FitToPagesWide = 1
                ws_pdf.PageSetup.FitToPagesTall = 1
                ws_pdf.PageSetup.Orientation = 2  # 2 = xlLandscape
                
                # Gera o nome do arquivo PDF e exporta
                pdf_nome = os.path.join(pasta_output, f"{nome_base}_ajustado.pdf")
                wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_nome))
                wb_pdf.Close(False)
                excel.Quit()
                print(f"PDF gerado: {pdf_nome}")
            except Exception as e:
                print(f"Erro ao converter para PDF: {e}")
        else:
            print("Conversão para PDF não disponível (win32com não instalado ou ambiente não Windows).")
        
    print("Processamento concluído!")
